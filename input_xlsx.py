# input_xlsx_no_argparse.py
# -*- coding: utf-8 -*-
"""
Excel から下記3種を JSON 化（argparse 不使用）
- 自動掲載シート: 0列目=項目名, 1列目以降=案件データ列 → 各列を1レコードに縦→横変換
- 自動返信シート: B1 のメッセージ
- 設定シート: A列=キー / B列=値 の辞書

使い方（最低限）:
    python input_xlsx_no_argparse.py input.xlsx extracted.json

オプション（環境変数で指定）:
    NORMALIZE_CHECKBOX=1   # ☑/☐ を True/False に正規化
    EXEC_ONLY=1            # 実行指示が☑の列だけ抽出
    POST_SHEET=自動掲載     # シート名を明示指定（未指定なら推測）
    REPLY_SHEET=自動返信
    SETTINGS_SHEET=設定
"""
from pathlib import Path
import os
import json
import pandas as pd
from openpyxl import load_workbook


def find_sheet_like(sheet_names, keywords):
    for name in sheet_names:
        if any(k in name for k in keywords):
            return name
    return None


def read_auto_post_records(xlsx_path, sheet_name, normalize_checkbox=False, start_row_auto=True):
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None, dtype=object)

    # 0列目の最初の非NaN行から「項目名が並ぶ」想定
    first_key_row = df[0].first_valid_index() if start_row_auto else 0
    if first_key_row is None:
        return []

    key_value_rows = df.iloc[first_key_row:, :].reset_index(drop=True)
    keys = key_value_rows.iloc[:, 0].tolist()
    max_col = key_value_rows.shape[1] - 1

    def _check_to_bool(v):
        if isinstance(v, str):
            if "☑" in v:
                return True
            if "☐" in v:
                return False
        return v

    records = []
    for col_idx in range(1, max_col + 1):
        col_series = key_value_rows.iloc[:, col_idx]
        if col_series.dropna().empty:
            continue

        rec = {"__col": col_idx}  # 元の列番号（デバッグ・追跡用）
        for k, v in zip(keys, col_series.tolist()):
            if pd.isna(k):
                continue
            key = str(k).strip()
            val = None if pd.isna(v) else v
            if normalize_checkbox:
                val = _check_to_bool(val)
            rec[key] = val

        records.append(rec)

    return records


def read_auto_reply_message(xlsx_path, sheet_name):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    return ws["B1"].value


def read_settings(xlsx_path, sheet_name):
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None, usecols=[0, 1], dtype=object)
    df = df.dropna(how="all")
    if df.empty:
        return {}

    def _is_headerish(a, b):
        a = "" if pd.isna(a) else str(a)
        b = "" if pd.isna(b) else str(b)
        return any(w in a for w in ["設定", "キー", "項目", "Key"]) or any(w in b for w in ["値", "Value"])

    if _is_headerish(df.iloc[0, 0], df.iloc[0, 1]):
        df = df.iloc[1:, :]

    out = {}
    for _, row in df.iterrows():
        k = row.iloc[0]
        if pd.isna(k):
            continue
        v = None if len(row) < 2 or pd.isna(row.iloc[1]) else row.iloc[1]
        out[str(k).strip()] = v
    return out


def get_input_json(
    excel_path: str,
    *,
    out: str | None = None,
    post_sheet: str | None = None,
    reply_sheet: str | None = None,
    settings_sheet: str | None = None,
    normalize_checkbox: bool = False,
    exec_only: bool = False,
):
    """プログラムから呼び出す用（argparse不使用）"""
    xlsx_path = Path(excel_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excelが見つかりません: {xlsx_path}")

    xls = pd.ExcelFile(xlsx_path)
    sheet_names = xls.sheet_names

    post_sheet = post_sheet or find_sheet_like(sheet_names, ["自動掲載", "掲載"])
    reply_sheet = reply_sheet or find_sheet_like(sheet_names, ["自動返信", "返信"])
    settings_sheet = settings_sheet or find_sheet_like(sheet_names, ["設定", "config", "Config"])

    auto_posts = []
    if post_sheet:
        auto_posts = read_auto_post_records(
            xlsx_path, post_sheet, normalize_checkbox=normalize_checkbox
        )
        if exec_only:
            def is_exec(rec):
                v = rec.get("実行指示")
                return bool(v) if normalize_checkbox else (isinstance(v, str) and ("☑" in v or "実行する" in v))
            auto_posts = [r for r in auto_posts if is_exec(r)]

    auto_reply_message = read_auto_reply_message(xlsx_path, reply_sheet) if reply_sheet else None
    settings = read_settings(xlsx_path, settings_sheet) if settings_sheet else {}

    bundle = {
        "auto_posts": auto_posts,
        "auto_reply_message": auto_reply_message,
        "settings": settings,
        "_meta": {
            "sheet_names": sheet_names,
            "detected_sheets": {
                "auto_post_sheet": post_sheet,
                "auto_reply_sheet": reply_sheet,
                "settings_sheet": settings_sheet,
            }
        }
    }

    if out:
        Path(out).write_text(json.dumps(bundle, ensure_ascii=False, indent=2), encoding="utf-8")

    return bundle


if __name__ == "__main__":
    import sys

    # 最低限の引数処理（argparse不使用）
    excel = sys.argv[1] if len(sys.argv) >= 2 else "input.xlsx"
    out = sys.argv[2] if len(sys.argv) >= 3 else "extracted.json"

    # 環境変数でオプション指定
    normalize_checkbox = os.getenv("NORMALIZE_CHECKBOX", "0").lower() in ("1", "true", "yes", "on")
    exec_only = os.getenv("EXEC_ONLY", "0").lower() in ("1", "true", "yes", "on")
    post_sheet = os.getenv("POST_SHEET") or None
    reply_sheet = os.getenv("REPLY_SHEET") or None
    settings_sheet = os.getenv("SETTINGS_SHEET") or None

    data = get_input_json(
        excel,
        out=out,
        post_sheet=post_sheet,
        reply_sheet=reply_sheet,
        settings_sheet=settings_sheet,
        normalize_checkbox=normalize_checkbox,
        exec_only=exec_only,
    )
    print(f"JSONを書き出しました: {Path(out).resolve()}")
